
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.axes
import matplotlib
from django.conf import settings
import matplotlib.patches as patches
import numpy as np
from sklearn.model_selection import train_test_split
from imblearn.over_sampling import RandomOverSampler, SMOTE
from sklearn.svm import SVC
from sklearn import linear_model
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import VotingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.preprocessing import OneHotEncoder
from xgboost import XGBClassifier
from xgboost import XGBRegressor
import lightgbm as lgb
import seaborn as sns
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import accuracy_score, classification_report, recall_score, precision_score, f1_score

#Excel de datos------------------------------------------------------------
datos = pd.read_excel('https://github.com/karique01/bomberospdfs/blob/master/dataset.xlsx?raw=true')##C:\Users\Data ##C:/Users/kvera/Documents/Proyectos/ml/dataset.xlsx

#region Modelo entrenamiento

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='PesoBebé')
# plt.show()

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='HistorialAborto')
# plt.show()

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='VisitaPrenatal')
# plt.show()

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='Paridad')
# plt.show()

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='ModoEntrega')
# plt.show()

# sns.countplot(data=datos, x='PuntajeApgarMin1',hue='EdadGestacional')
# plt.show()

#Eliminar Columna que no se evaluará (ID)
#datos = datos.drop(columns=['ID'])
#datos = datos.drop(columns=['LLantoBebé'])
#datos = datos.drop(columns=['LugarDeNacimiento'])

#Visualizar Tipo de dato
#datos.info()

#Revisar que no falten datos en las columnas
#print(datos.isna().sum())

#Cabecera
#print(datos.all)
#print(datos.head())

#CATEGORIA DE DATOS----------------------------------------------------------
#Tratar Datos Paridad
#print(datos['Paridad'].unique())
valores_cross1 = {"Paridad":  {'primiparious':0, 'multiparious':1}}
datos.replace(valores_cross1, inplace=True)
#print(datos.head())

# #Tratar Datos VisitaPrenatal
# #print(datos['VisitaPrenatal'].unique())
valores_cross2 = {"VisitaPrenatal":  {'No':0, 'Yes':1}}
datos.replace(valores_cross2, inplace=True)
# #print(datos.head())

# #Tratar ModoEntrega
# #print(datos['ModoEntrega'].unique()) - 1
# valores_cross4 = {"ModoEntrega":  {'SVD':0, 'AVD':1, 'C/S':2}}
# datos.replace(valores_cross4, inplace=True)
# #print(datos.head())

# #Tratar CondiciónDeParto
# #print(datos['CondiciónDeParto'].unique()) - 2
# valores_cross5 = {"CondiciónDeParto":  {'Without labor':0, 'Spontanious':1, 'Induced':2}}
# datos.replace(valores_cross5, inplace=True)
# #print(datos.head())

# #Tratar DuraciónParto
# #print(datos['DuraciónParto'].unique()) - 3
# valores_cross6 = {"DuraciónParto":  {'Preciptated':0, 'Normal':1, 'Prolonged':2}}
# datos.replace(valores_cross6, inplace=True)
# #print(datos.head())

#Tratar Duración_RoturaDeMembrana
#print(datos['Duración_RoturaDeMembrana'].unique())
valores_cross7 = {"Duración_RoturaDeMembrana":  {'< 18 hrs':0, '> 18hrs':1}}
datos.replace(valores_cross7, inplace=True)
#print(datos.head())

#Tratar Tipo de embarazo
#print(datos['Tipo de embarazo'].unique())
valores_cross8 = {"Tipo de embarazo":  {'singlton':0, 'multiple':1}}
datos.replace(valores_cross8, inplace=True)
#print(datos['Tipo de embarazo'].unique())

#Tratar Sufrimiento Fetal
valores_cross9 = {"Sufrimiento Fetal":  {'No':0, 'Yes':1}}
datos.replace(valores_cross9, inplace=True)

#Tratar ProblemaCordónUmbilical
valores_cross10 = {"ProblemaCordónUmbilical":  {'No':0, 'Yes':1}}
datos.replace(valores_cross10, inplace=True)

#Tratar PocoFluidoAmniótico
valores_cross11 = {"PocoFluidoAmniótico":  {'No':0, 'Yes':1}}
datos.replace(valores_cross11, inplace=True)

#Tratar HistorialAborto
valores_cross12 = {"HistorialAborto":  {'No':0, 'Yes':1}}
datos.replace(valores_cross12, inplace=True)

#Tratar Hipertensión
valores_cross13 = {"Hipertensión":  {'No':0, 'Yes':1}}
datos.replace(valores_cross13, inplace=True)

#Tratar HemorragiaAnteparto
valores_cross14 = {"HemorragiaAnteparto":  {'No':0, 'Yes':1}}
datos.replace(valores_cross14, inplace=True)

#Tratar Sexo 
valores_cross15 = {"Sexo":  {'Female':0, 'Male':1}}
datos.replace(valores_cross15, inplace=True)

# #Tratar EdadGestacional
# #print(datos['EdadGestacional'].unique()) 4
# valores_cross16 = {"EdadGestacional":  {'Preterm baby':0, 'Term baby':1, 'Postterm':2}}
# datos.replace(valores_cross16, inplace=True)
# #print(datos['EdadGestacional'].unique())

#Tratar PuntajeApgarMin1
#print(datos['PuntajeApgarMin1'].unique())
valores_cross18 = {"PuntajeApgarMin1":  {'>= 7':0, '< 7':1}}
datos.replace(valores_cross18, inplace=True)
#print(datos['PuntajeApgarMin1'].unique())

# #Tratar EdadMadre
# #print(datos['EdadMadre'].unique()) - 5
# valores_cross19 = {"EdadMadre":  {'< 20 years':0, '20-34 years':1, '>35 years':2}}
# datos.replace(valores_cross19, inplace=True)
# #print(datos['EdadMadre'].unique())

# #Tratar PesoBebé
# #print(datos['PesoBebé'].unique()) - 6
# valores_cross20 = {"PesoBebé":  {'Lowbirth weight':0, 'Normal':1, 'Big baby':2}}
# datos.replace(valores_cross20, inplace=True)
# #print(datos['PesoBebé'].unique())

#IMAGENES------------------------------------------------------------------------------
# sns.histplot(data=datos, x='PuntajeApgarMin1',hue='PesoBebé',multiple='stack')
# plt.show()

# sns.histplot(data=datos, x='PuntajeApgarMin1',hue='HistorialAborto',multiple='stack')
# plt.show()

# sns.histplot(data=datos, x='PuntajeApgarMin1',hue='VisitaPrenatal',multiple='stack')
# plt.show()

# sns.histplot(data=datos, x='PuntajeApgarMin1',hue='Paridad',multiple='stack')
# plt.show()

# col_cat = [['EdadMadre', 'Paridad', 'VisitaPrenatal', 'ModoEntrega',
#        'CondiciónDeParto', 'DuraciónParto', 'Duración_RoturaDeMembrana',
#        'Tipo de embarazo', 'Sufrimiento Fetal', 'ProblemaCordónUmbilical',
#        'PocoFluidoAmniótico', 'HistorialAborto', 'Hipertensión',
#        'HemorragiaAnteparto', 'Sexo', 'PesoBebé', 'EdadGestacional']]

# fig, ax = plt.subplots(nrows=18,ncols=1)
# fig.subplots_adjust(hspace=0.5)

# for i, col in enumerate (col_cat):
#       sns.countplot(x=col, data=datos,ax=ax[i])
#       ax[i].set_title(col)
#       ax[i].set_xticklabels(ax[i].get_xticklabels(),rotation=30)

#AQUI-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#transformación de datos 
transform_datos = OneHotEncoder(handle_unknown='ignore', sparse_output=False).set_output(transform='pandas')
nuevas_cols = transform_datos.fit_transform(datos[['ModoEntrega', 'CondiciónDeParto', 'DuraciónParto', 'EdadGestacional', 'EdadMadre', 'PesoBebé']])
nueva_data = pd.concat([datos, nuevas_cols],axis=1).drop(columns=['ModoEntrega', 'CondiciónDeParto', 'DuraciónParto', 'EdadGestacional', 'EdadMadre', 'PesoBebé','CondiciónDeParto_Without labor'])
#print(nueva_data.head(10))
#nueva_data.info()

#Revisar que no falten datos en las columnas
#print(nueva_data.isna().sum())
#Selección de variables Stepwise Regresion

#datos de registros sin contar columna apgar
X_var = nueva_data.drop('PuntajeApgarMin1',axis=1)

# #datos de la columna apagar
y_var = nueva_data.PuntajeApgarMin1

#Análisis exploratorio de datos
#sns.barplot(data=nueva_data, x='PuntajeApgarMin1',y='Paridad')
#plt.show()

#Stepwise Regression backward
sfs = SequentialFeatureSelector(linear_model.LogisticRegression(), n_features_to_select=5, direction='forward', scoring='accuracy', cv=None)

variables = sfs.fit(X_var, y_var)
print(variables.get_feature_names_out())



#Columnas seleccionadas - ['Paridad' 'VisitaPrenatal' 'Duración_RoturaDeMembrana' 'ModoEntrega_C/S' 'EdadMadre_20-34 years']
datos_selected1 = nueva_data[variables.get_feature_names_out()]
datos_selected2 = nueva_data[['PuntajeApgarMin1']]
#Visualizar Tipo de dato
#datos_selected1.info()
#datos_selected2.info()

datos_selected = pd.concat([datos_selected1,datos_selected2],axis=1)
#Revisar que no falten datos en las columnas
#print(datos_selected.isna().sum())
#print(datos_selected.info())
#print(datos_selected.head(10))

#print(datos_selected.PuntajeApgarMin1.value_counts(normalize=True)*100)

#Separación Train, Test y Validation
# #datos de registros sin contar columna apgar
X_datos = datos_selected.drop('PuntajeApgarMin1',axis=1)

# #datos de la columna apagar
y_datos = datos_selected.PuntajeApgarMin1

#print(y_datos.value_counts(normalize=True)*100)

# # Separamos en X e Y las variables - 80% para ser entrenamiento y testeo (X_data_eval , y_data_eval) - 20% para validación (X_data_validar , y_data_validar)
X_data_eval, X_data_validar, y_data_eval, y_data_validar = train_test_split(X_datos, y_datos, test_size=0.2,random_state=2023, stratify = y_datos)
#print(y_data_eval.value_counts(normalize=True)*100)
#print(y_data_validar.value_counts(normalize=True)*100)

# # Separamos en X e Y las variables - 80% para entrenamiento (X_train_eval , y_train_eval) - 20% para test (X_test_eval , y_test_eval)
X_train_eval_bal, X_test_eval, y_train_eval_bal, y_test_eval = train_test_split(X_data_eval, y_data_eval, test_size=0.2,random_state=2023, stratify = y_data_eval)
#print(y_train_eval_bal.value_counts(normalize=True)*100)
#print(y_test_eval.value_counts(normalize=True)*100)

#print(X_train_eval)
#print(y_train_eval)

#Balanceo de datos de entrenamiento---------------------------------------------------
# ros = RandomOverSampler()
smote = SMOTE()

X_train_eval, y_train_eval = smote.fit_resample(X_train_eval_bal, y_train_eval_bal)

# print(y_data_eval.value_counts(normalize=True)*100)
# print(y_data_validar.value_counts(normalize=True)*100)
# print(y_train_eval_bal.value_counts(normalize=True)*100)
# print(y_test_eval.value_counts(normalize=True)*100)

# print(y_data_eval.value_counts())
# print(y_data_validar.value_counts())
# print(y_train_eval.value_counts())
# print(y_test_eval.value_counts())

#datos de registros sin contar columna apgar
#X_datos = datos_selected.drop('PuntajeApgarMin1',axis=1)

#datos de la columna apagar
#y_datos = datos_selected.PuntajeApgarMin1

# #PASAR DATOS BALANCEADOS A EXCEL
# datos_bal, target_bal = ros.fit_resample(X_datos, y_datos)

# #DESCARGAR DATOS A EXCEL--------------------------------------------
# #datos_bal.to_excel("Datos_final.xlsx")
# #target_bal.to_excel("asfixiados.xlsx")

# #datos.info()

# #Proporcion de la variable obejtivo (ASFIXIA)
# #plt.figure(figsize=(7,7))

# #labels, counts = np.unique(target_bal, return_counts=True)
# #plt.pie(counts, autopct='%1.1f%%',labels=labels)
# #plt.legend({'Asfixia','No Asfixia'})
# #plt.title('Proporcion de Asfixia en neonatos')

# #GRAFICA DE NEONATOS CON O SIN ASFIXIA
# #plt.show()
# #Conteo de datos (1:Asfixia,2:No Asfixia)
# #print(target_bal.value_counts())

# #-------------------------------------------------------------------
# #Evaluación

# # Entrenamiento de datos
# # Separamos en X e Y las variables

# X_train_eval, X_test_eval, y_train_eval, y_test_eval = train_test_split(datos_bal, target_bal, test_size=0.2)

#RANDOM FOREST
print('VALORES - RANDOM FOREST--------')
#hiperparámetros Bosques Aleatorios

#Definir los hiperparámetros y sus posibles valores
param_grid_rnd = {
      'n_estimators': [10,25,50],
      'max_depth' : [5,10,15],
      'criterion' : ['gini', 'entropy', 'log_loss'],
      'min_samples_split': [2,4,6],
      'min_samples_leaf': [1,2,4]
  }

#Crear el objeto GridSearchCV
grid_search_rnd = GridSearchCV(estimator=RandomForestClassifier(), param_grid=param_grid_rnd, cv=5, scoring='accuracy', n_jobs=-1)

#Ajustar el modelo con GridSearchCV
grid_search_rnd.fit(X_train_eval, y_train_eval)

model_rnd = RandomForestClassifier().set_params(**grid_search_rnd.best_params_)
model_rnd.fit(X_train_eval, y_train_eval)
rnd_train = model_rnd.predict(X_train_eval)
rnd_eval = model_rnd.predict(X_test_eval)
rnd_valid = model_rnd.predict(X_data_validar)

#metrica RANDOM FOREST - train
print('metrica RANDOM FOREST - train-----------------------')
print(classification_report(y_train_eval, rnd_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,rnd_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,rnd_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,rnd_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,rnd_train)))

#metrica RANDOM FOREST - test
print('metrica RANDOM FOREST - test-----------------------')
print(classification_report(y_test_eval,rnd_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,rnd_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,rnd_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,rnd_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,rnd_eval)))

#metrica RANDOM FOREST - Validación
print('metrica RANDOM FOREST - validación-----------------------')
print(classification_report(y_data_validar, rnd_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,rnd_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,rnd_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,rnd_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,rnd_valid)))

#Regresion logistica
print('VALORES - REGRESION LOGISTICA------------')
#hiperparámetros REGRESION LOGISTICA

#Definir los hiperparámetros y sus posibles valores
param_grid_lr = {
      'penalty' : [ 'l2'],
      'C': np.logspace(-4, 4, 20),
      'solver' : ['lbfgs', 'liblinear', 'newton-cg', 'sag', 'saga'],
      'max_iter' : [100, 1000, 2500, 5000]
  }

#Crear el objeto GridSearchCV
grid_search_lr = GridSearchCV(estimator=LogisticRegression(), param_grid=param_grid_lr, cv=5, scoring='accuracy')

#Ajustar el modelo con GridSearchCV
grid_search_lr.fit(X_train_eval, y_train_eval)

model_lr = LogisticRegression().set_params(**grid_search_lr.best_params_)
model_lr.fit(X_train_eval, y_train_eval)
lr_train = model_lr.predict(X_train_eval)
lr_eval = model_lr.predict(X_test_eval)
lr_valid = model_lr.predict(X_data_validar)

#metrica REGRESION LOGISTICA - train
print('metrica REGRESION LOGISTICA - train-----------------------')
print(classification_report(y_train_eval, lr_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,lr_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,lr_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,lr_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,lr_train)))

#metrica REGRESION LOGISTICA - test
print('metrica REGRESION LOGISTICA - test-----------------------')
print(classification_report(y_test_eval,lr_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,lr_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,lr_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,lr_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,lr_eval)))

#metrica REGRESION LOGISTICA - Validación
print('metrica REGRESION LOGISTICA - validación-----------------------')
print(classification_report(y_data_validar, lr_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,lr_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,lr_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,lr_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,lr_valid)))

#XGBOOST
print('VALORES - XGBOOST---------------')
#hiperparámetros XGBOOST

#Definir los hiperparámetros y sus posibles valores

param_grid_xgb = {
      'n_estimators' : [ 100, 200, 500],
      'max_depth' : [3, 6, 9],
      'gamma' : [0.01, 0.1],
      'learning_rate' : [0.001, 0.01, 0.1, 1]
  }

#Crear el objeto GridSearchCV
grid_search_xgb = GridSearchCV(estimator=XGBClassifier(), param_grid=param_grid_xgb, scoring= ['r2', 'neg_root_mean_squared_error'], refit='r2', cv=5)

#Ajustar el modelo con GridSearchCV
grid_search_xgb.fit(X_train_eval, y_train_eval)

model_xgb = XGBClassifier().set_params(**grid_search_xgb.best_params_) 
model_xgb.fit(X_train_eval, y_train_eval)
xgb_train = model_xgb.predict(X_train_eval)
xgb_eval = model_xgb.predict(X_test_eval)
xgb_valid = model_xgb.predict(X_data_validar)

#metrica XGBOOST - train
print('metrica XGBOOST - train-----------------------')
print(classification_report(y_train_eval, xgb_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,xgb_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,xgb_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,xgb_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,xgb_train)))

#metrica XGBOOST - test
print('metrica XGBOOST - test-----------------------')
print(classification_report(y_test_eval,xgb_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,xgb_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,xgb_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,xgb_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,xgb_eval)))

#metrica XGBOOST - Validación
print('metrica XGBOOST - validación-----------------------')
print(classification_report(y_data_validar, xgb_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,xgb_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,xgb_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,xgb_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,xgb_valid)))

#KNeighborsClassifier
print('VALORES - KNeighborsClassifier---------------')
#hiperparámetros KNeighborsClassifier
param_grid_knn = {
      'n_neighbors': range(1, 30, 2),
      'weights' : ['uniform', 'distance'],
      'metric' : ['euclidean', 'manhattan', 'minkowski'],
      'leaf_size' : range(1, 50, 5)
}

#Crear el objeto GridSearchCV
grid_search_knn = GridSearchCV(estimator=KNeighborsClassifier(), param_grid=param_grid_knn, n_jobs=-1,cv=5,scoring='accuracy')

#Ajustar el modelo con GridSearchCV
grid_search_knn.fit(X_train_eval, y_train_eval)

model_knn = KNeighborsClassifier().set_params(**grid_search_knn.best_params_)
model_knn.fit(X_train_eval, y_train_eval)
knn_train = model_knn.predict(X_train_eval)
knn_eval = model_knn.predict(X_test_eval)
knn_valid = model_knn.predict(X_data_validar)

#metrica KNeighborsClassifier - train
print('metrica KNeighborsClassifier - train-----------------------')
print(classification_report(y_train_eval, knn_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,knn_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,knn_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,knn_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,knn_train)))

#metrica KNeighborsClassifier - test
print('metrica KNeighborsClassifier - test-----------------------')
print(classification_report(y_test_eval,knn_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,knn_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,knn_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,knn_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,knn_eval)))

#metrica KNeighborsClassifier - Validación
print('metrica KNeighborsClassifier - validación-----------------------')
print(classification_report(y_data_validar, knn_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,knn_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,knn_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,knn_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,knn_valid)))

#DecisionTreeClassifier
print('VALORES - DecisionTreeClassifier---------------')
#hiperparámetros DecisionTreeClassifier
param_grid_DT = {
      'max_depth': [3, 4, None],
      'min_samples_leaf' : [1 , 2 , 3],
      'criterion' : ['gini', 'entropy']
}

#Crear el objeto GridSearchCV
grid_search_DT = GridSearchCV(estimator=DecisionTreeClassifier(), param_grid=param_grid_DT, cv = 5, scoring = 'accuracy')

#Ajustar el modelo con GridSearchCV
grid_search_DT.fit(X_train_eval, y_train_eval)

model_DT = DecisionTreeClassifier().set_params(**grid_search_DT.best_params_)
model_DT.fit(X_train_eval, y_train_eval)
DT_train = model_DT.predict(X_train_eval)
DT_eval = model_DT.predict(X_test_eval)
DT_valid = model_DT.predict(X_data_validar)

#metrica DecisionTreeClassifier - train
print('metrica DecisionTreeClassifier - train-----------------------')
print(classification_report(y_train_eval, DT_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,DT_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,DT_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,DT_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,DT_train)))

#metrica DecisionTreeClassifier - test
print('metrica DecisionTreeClassifier - test-----------------------')
print(classification_report(y_test_eval,DT_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,DT_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,DT_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,DT_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,DT_eval)))

#metrica DecisionTreeClassifier - Validación
print('metrica DecisionTreeClassifier - validación-----------------------')
print(classification_report(y_data_validar, DT_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,DT_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,DT_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,DT_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,DT_valid)))

#Support vector machines(Support Vector Classification) - Máquinas de vectores de soporte(Clasificación de vectores de soporte)
print('VALORES - Support Vector Classification---------------')
#hiperparámetros Support Vector Classification
param_grid_svm = {
      'C': [1.0, 6.0, 10.0],
      'kernel' : ['rbf', 'poly'],
      'gamma' : ['scale', 'auto'],
      'probability' : [True]
}

#Crear el objeto GridSearchCV
grid_search_svm = GridSearchCV(estimator=SVC(), param_grid=param_grid_svm, cv = 5, scoring = 'accuracy', n_jobs=-1)

#Ajustar el modelo con GridSearchCV
grid_search_svm.fit(X_train_eval, y_train_eval)


model_svm = SVC().set_params(**grid_search_svm.best_params_)
model_svm.fit(X_train_eval, y_train_eval)
svm_train = model_svm.predict(X_train_eval)
svm_eval = model_svm.predict(X_test_eval)
svm_valid = model_svm.predict(X_data_validar)

#metrica Support Vector Classification - train
print('metrica Support Vector Classification - train-----------------------')
print(classification_report(y_train_eval, svm_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,svm_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,svm_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,svm_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,svm_train)))

#metrica Support Vector Classification - test
print('metrica Support Vector Classification - test-----------------------')
print(classification_report(y_test_eval,svm_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,svm_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,svm_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,svm_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,svm_eval)))

#metrica Support Vector Classification - Validación
print('metrica Support Vector Classification - validación-----------------------')
print(classification_report(y_data_validar, svm_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,svm_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,svm_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,svm_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,svm_valid)))

#neural network(Multi-layer Perceptron classifier) - red neuronal(Clasificador de perceptrones multicapa)
print('VALORES - Multi-layer Perceptron classifier---------------')
#hiperparámetros Multi-layer Perceptron classifier
param_grid_MLP = {
      'hidden_layer_sizes': [5],
      'activation' : ['identity', 'logistic', 'tanh', 'relu'],
      'solver' : ['lbfgs'],#, 'sgd', 'adam'],
      'learning_rate' : ['constant', 'invscaling', 'adaptive'],
      'max_iter' : [1000]
}

#Crear el objeto GridSearchCV
grid_search_mlp = GridSearchCV(estimator=MLPClassifier(), param_grid=param_grid_MLP, cv = 5, scoring = 'accuracy')

#Ajustar el modelo con GridSearchCV
grid_search_mlp.fit(X_train_eval, y_train_eval)


model_mlp = MLPClassifier().set_params(**grid_search_mlp.best_params_)
model_mlp.fit(X_train_eval, y_train_eval)
mlp_train = model_mlp.predict(X_train_eval)
mlp_eval = model_mlp.predict(X_test_eval)
mlp_valid = model_mlp.predict(X_data_validar)

#metrica Multi-layer - train
print('metrica Multi-layer - train-----------------------')
print(classification_report(y_train_eval, mlp_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,mlp_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,mlp_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,mlp_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,mlp_train)))

#metrica Multi-layer - test
print('metrica Multi-layer - test-----------------------')
print(classification_report(y_test_eval,mlp_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,mlp_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,mlp_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,mlp_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,mlp_eval)))

#metrica Multi-layer - Validación
print('metrica Multi-layer - validación-----------------------')
print(classification_report(y_data_validar, mlp_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,mlp_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,mlp_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,mlp_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,mlp_valid)))

#VotingClassifier
print('VotingClassifier-----------------------')
voting_model = VotingClassifier(estimators=[
      ('model_lr',model_lr),
      ('model_svm',model_svm),
      ('model_rnd',model_rnd),
      ('model_knn',model_knn),
      ('model_xgb',model_xgb),
      ('model_DT',model_DT),
      ('model_MLP',model_mlp)
  ],voting='soft')



voting_model.fit(X_train_eval, y_train_eval)
voting_train = voting_model.predict(X_train_eval)
voting_eval = voting_model.predict(X_test_eval)
voting_valid = voting_model.predict(X_data_validar)

#metrica Voting - train
print('metrica Voting - train-----------------------')
print(classification_report(y_train_eval, voting_train))
print('Accuracy Score : ' + str(accuracy_score(y_train_eval,voting_train)))
print('Precision Score : ' + str(precision_score(y_train_eval,voting_train)))
print('Recall Score : ' + str(recall_score(y_train_eval,voting_train)))
print('F1 Score : ' + str(f1_score(y_train_eval,voting_train)))
#print(voting_eval)

#metrica Voting - test
print('metrica Voting - test-----------------------')
print(classification_report(y_test_eval, voting_eval))
print('Accuracy Score : ' + str(accuracy_score(y_test_eval,voting_eval)))
print('Precision Score : ' + str(precision_score(y_test_eval,voting_eval)))
print('Recall Score : ' + str(recall_score(y_test_eval,voting_eval)))
print('F1 Score : ' + str(f1_score(y_test_eval,voting_eval)))
#print(voting_eval)

#metrica Voting - Validación
print('metrica Voting - validación-----------------------')
print(classification_report(y_data_validar, voting_valid))
print('Accuracy Score : ' + str(accuracy_score(y_data_validar,voting_valid)))
print('Precision Score : ' + str(precision_score(y_data_validar,voting_valid)))
print('Recall Score : ' + str(recall_score(y_data_validar,voting_valid)))
print('F1 Score : ' + str(f1_score(y_data_validar,voting_valid)))

#endregion


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, paridad, visita_prenatal, duracion_rotura_membrana, parto_cesarea, edad_madre_veinte_treintacuatro, foto_empleado, salario_empleado, modo_entrega, condicion_parto, duracion_parto, tipo_embarazo, sufrimiento_fetal, problema_cordon_umbilical, poco_fluido_amniotico, historial_aborto, hipertension, hemorragia_anteparto, peso_bebe, edad_gestacional) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                           dataForm['paridad'], dataForm['visita_prenatal'], dataForm['duracion_rotura_membrana'], 
                           dataForm['parto_cesarea'], dataForm['edad_madre_veinte_treintacuatro'], 
                           result_foto_perfil, salario_entero,
                           dataForm['modo_entrega'], dataForm['condicion_parto'],
                           dataForm['duracion_parto'], dataForm['tipo_embarazo'],
                           dataForm['sufrimiento_fetal'], dataForm['problema_cordon_umbilical'],
                           dataForm['poco_fluido_amniotico'], dataForm['historial_aborto'],
                           dataForm['hipertension'], dataForm['hemorragia_anteparto'],
                           dataForm['peso_bebe'], dataForm['edad_gestacional'])
                
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro,
                        e.paridad,
                        e.visita_prenatal,
                        e.duracion_rotura_membrana,
                        e.parto_cesarea,
                        e.edad_madre_veinte_treintacuatro,
                        e.modo_entrega,
                        e.condicion_parto,
                        e.duracion_parto,
                        e.tipo_embarazo,
                        e.sufrimiento_fetal,
                        e.problema_cordon_umbilical,
                        e.poco_fluido_amniotico,
                        e.historial_aborto,
                        e.hipertension,
                        e.hemorragia_anteparto,
                        e.peso_bebe,
                        e.edad_gestacional,
                        '' as asfixia
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado ,
                        e.nombre_empleado , 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        e.modo_entrega,
                        e.condicion_parto,
                        e.duracion_parto,
                        e.tipo_embarazo,
                        e.sufrimiento_fetal,
                        e.problema_cordon_umbilical,
                        e.poco_fluido_amniotico,
                        e.historial_aborto,
                        e.hipertension,
                        e.hemorragia_anteparto,
                        e.peso_bebe,
                        e.edad_gestacional,
                        e.paridad,
                        e.visita_prenatal,
                        e.duracion_rotura_membrana,
                        e.edad_madre_veinte_treintacuatro,
                        e.parto_cesarea,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Paridad", "Visita Prenatal", "Duracion de Rotura de Menbrana",
                     "Modo entrega", "Condición de Parto", "Duración parto", "Tipo de embarazo", 
                     "Sufrimiento Fetal", "Problema de cordón umbilical", "Poco Fluido Amniotico", "Problema de cordón umbilical", 
                     "Hipertensión", "Hemorragia Anteparto", "Peso Bebé", "Edad Gestacional", 
                     "Edad de la madre [20-34] años", "Parto por cesárea", 
                     "Peso", 
                     "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    #formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']


        paridad = 'Primípara' if registro['paridad'] == 0 else 'Multípara'
        visita_prenatal = 'No' if registro['visita_prenatal'] == 0 else 'Si'
        duracion_rotura_membrana = '< 18 hrs' if registro['duracion_rotura_membrana'] == 0 else '> 18hrs'
        modo_entrega = 'AVD' if registro['modo_entrega'] == 1 else 'C/S' if registro['modo_entrega'] == 2 else 'SVD'
        condicion_parto = 'Inducida' if registro['condicion_parto'] == 1 else 'Espontanea' if registro['condicion_parto'] == 2 else 'Sin Labor'
        duracion_parto = 'Normal' if registro['duracion_parto'] == 1 else 'Precipitado' if registro['duracion_parto'] == 2 else 'Prolongado'
        tipo_embarazo = 'Multiple' if registro['tipo_embarazo'] == 1 else 'Único'
        sufrimiento_fetal = 'No' if registro['sufrimiento_fetal'] == 0 else 'Si'
        problema_cordon_umbilical = 'No' if registro['problema_cordon_umbilical'] == 0 else 'Si'
        poco_fluido_amniotico = 'No' if registro['poco_fluido_amniotico'] == 0 else 'Si'
        historial_aborto = 'No' if registro['historial_aborto'] == 0 else 'Si'
        hipertension = 'No' if registro['hipertension'] == 0 else 'Si'
        hemorragia_anteparto = 'No' if registro['hemorragia_anteparto'] == 0 else 'Si'
        peso_bebe = 'Bebé Grande' if registro['peso_bebe'] == 1 else 'Peso Ligero' if registro['peso_bebe'] == 2 else 'Normal'
        edad_gestacional = 'Post Termino' if registro['edad_gestacional'] == 1 else 'Pre Termino' if registro['edad_gestacional'] == 2 else 'Bebé a termino'
        edad_madre_veinte_treintacuatro = 'No' if registro['edad_madre_veinte_treintacuatro'] == 0 else 'Si'
        parto_cesarea = 'No' if registro['parto_cesarea'] == 0 else 'Si'


        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']
    
        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, 
                     paridad, visita_prenatal, duracion_rotura_membrana,
                     modo_entrega,condicion_parto,duracion_parto,tipo_embarazo,
                     sufrimiento_fetal,problema_cordon_umbilical,poco_fluido_amniotico,historial_aborto,
                     hipertension,hemorragia_anteparto,peso_bebe,edad_gestacional,
                     edad_madre_veinte_treintacuatro,parto_cesarea,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        # for fila_num in range(2, hoja.max_row + 1):
        #     columna = 7  # Columna G
        #     celda = hoja.cell(row=fila_num, column=columna)
        #     celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_pacientes_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado,
                            e.modo_entrega,
                            e.condicion_parto,
                            e.duracion_parto,
                            e.tipo_embarazo,
                            e.sufrimiento_fetal,
                            e.problema_cordon_umbilical,
                            e.poco_fluido_amniotico,
                            e.historial_aborto,
                            e.hipertension,
                            e.hemorragia_anteparto,
                            e.peso_bebe,
                            e.edad_gestacional
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                modo_entrega = data.form['modo_entrega']
                condicion_parto = data.form['condicion_parto']
                duracion_parto = data.form['duracion_parto']
                tipo_embarazo = data.form['tipo_embarazo']
                sufrimiento_fetal = data.form['sufrimiento_fetal']
                problema_cordon_umbilical = data.form['problema_cordon_umbilical']
                poco_fluido_amniotico = data.form['poco_fluido_amniotico']
                historial_aborto = data.form['historial_aborto']
                hipertension = data.form['hipertension']
                hemorragia_anteparto = data.form['hemorragia_anteparto']
                peso_bebe = data.form['peso_bebe']
                edad_gestacional = data.form['edad_gestacional']

                salario_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            salario_empleado = %s,
                            foto_empleado = %s,
                            modo_entrega = %s,
                            condicion_parto = %s,
                            duracion_parto = %s,
                            tipo_embarazo = %s,
                            sufrimiento_fetal = %s,
                            problema_cordon_umbilical = %s,
                            poco_fluido_amniotico = %s,
                            historial_aborto = %s,
                            hipertension = %s,
                            hemorragia_anteparto = %s,
                            peso_bebe = %s,
                            edad_gestacional = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              salario_empleado, fotoForm, 
                              modo_entrega, condicion_parto, duracion_parto, tipo_embarazo, sufrimiento_fetal, problema_cordon_umbilical, poco_fluido_amniotico, historial_aborto, hipertension, hemorragia_anteparto, peso_bebe, edad_gestacional,
                              id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            salario_empleado = %s,
                            modo_entrega = %s,
                            condicion_parto = %s,
                            duracion_parto = %s,
                            tipo_embarazo = %s,
                            sufrimiento_fetal = %s,
                            problema_cordon_umbilical = %s,
                            poco_fluido_amniotico = %s,
                            historial_aborto = %s,
                            hipertension = %s,
                            hemorragia_anteparto = %s,
                            peso_bebe = %s,
                            edad_gestacional = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              salario_empleado, 
                              modo_entrega, condicion_parto, duracion_parto, tipo_embarazo, sufrimiento_fetal, problema_cordon_umbilical, poco_fluido_amniotico, historial_aborto, hipertension, hemorragia_anteparto, peso_bebe, edad_gestacional,
                              id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
